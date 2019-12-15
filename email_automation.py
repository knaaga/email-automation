import imapclient
import imaplib
import smtplib
import pyzmail
import time
import pyprind
import re
import urllib.parse
from email.mime.multipart import MIMEMultipart
from openpyxl import Workbook


def login(imap_server, smtp_server, username, password):
    imapobj = imapclient.IMAPClient(imap_server, ssl=True)
    imapobj.login(username, password)
    smtpobj = smtplib.SMTP(smtp_server, 587)
    smtpobj.ehlo()
    smtpobj.starttls()
    smtpobj.login(username, password)
    print("Login successful....")
    return imapobj, smtpobj


def retrieve_emails(imapobj):
    print("Retrieving emails....")
    imaplib._MAXLINE = 10000000
    imapobj.select_folder('Inbox', readonly=True)
    UIDs = imapobj.search(['SINCE', '01-Aug-2019', 'BEFORE', '03-Aug-2019'])
    return UIDs


def categorize_emails(imapobj, UIDs):
    category = []
    for i in range(len(UIDs)):
        label_dict = imapobj.get_gmail_labels(UIDs[i])
        label = label_dict[UIDs[i]]
        if 'Starred' in str(label):
            category.append('Starred')
        elif 'Important' in str(label):
            category.append('Important')
        elif len(label) == 0:
            category.append('Inbox')
        else:
            category.append('Custom Label')
    return category


def email_attributes(imapobj, UIDs):
    print("Extracting email attributes....")
    from_addresses = []
    subjects = []
    dates = []
    days = []
    months = []
    years = []
    times = []
    sent_received = []
    unsub_links = []
    category = categorize_emails(imapobj, UIDs)
    email_data = []

    bar = pyprind.ProgBar(len(UIDs))

    for i in range(len(UIDs)):
        raw_message = imapobj.fetch(UIDs[i], ['BODY[]'])
        message = pyzmail.PyzMessage.factory(raw_message[UIDs[i]][b'BODY[]'])

        if message.get_address('from')[1] == username:
            full_date = message.get_decoded_header('date')
            sent_received.append('Sent')
        else:
            sent_received.append('Received')
            full_date = message.get_decoded_header('Received').split('\n')[1].strip(' ')

        from_addresses.append(message.get_address('from'))
        subjects.append(message.get_subject(''))
        unsub_link = message.get_decoded_header('List-Unsubscribe')
        if len(str(unsub_link)) > 0 and 'mailto' in unsub_link:
            unsub_link = unsub_link.split(',')
            unsub_links.append([unsub_link[idx] for idx, s in enumerate(unsub_link) if 'mailto' in s][0])
        else:
            unsub_links.append('No unsubscribe link found')

        day = full_date.split()[0].strip(',')
        date = full_date.split()[1]
        month = full_date.split()[2]
        year = full_date.split()[3]
        time2 = full_date.split()[4]

        days.append(day)
        dates.append(date)
        months.append(month)
        years.append(year)
        times.append(time2)

        time.sleep(0.01)
        bar.update()

    email_data.extend([dates, months, years, days, times, from_addresses, subjects, sent_received, category, unsub_links])
    return email_data


def write_to_excel(email_data):
    print("Writing data to excel")
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.cell(1, 1).value = "Date"
    ws.cell(1, 2).value = "Month"
    ws.cell(1, 3).value = "Year"
    ws.cell(1, 4).value = "Day"
    ws.cell(1, 5).value = "Time"
    ws.cell(1, 6).value = "From (Sender)"
    ws.cell(1, 7).value = "From (Email ID)"
    ws.cell(1, 8).value = "Subject"
    ws.cell(1, 9).value = "Sent/Received"
    ws.cell(1, 10).value = "Category"
    ws.cell(1, 11).value = "Unsubscribe Link"

    for i in range(len(UIDs)):
        ws.cell(row=i + 2, column=1).value = email_data[0][i]
        ws.cell(row=i + 2, column=2).value = email_data[1][i]
        ws.cell(row=i + 2, column=3).value = email_data[2][i]
        ws.cell(row=i + 2, column=4).value = email_data[3][i]
        ws.cell(row=i + 2, column=5).value = email_data[4][i]
        ws.cell(row=i + 2, column=6).value = email_data[5][i][0]
        ws.cell(row=i + 2, column=7).value = email_data[5][i][1]
        ws.cell(row=i + 2, column=8).value = str(email_data[6][i])
        ws.cell(row=i + 2, column=9).value = email_data[7][i]
        ws.cell(row=i + 2, column=10).value = email_data[8][i]
        ws.cell(row=i + 2, column=11).value = email_data[9][i]

    wb.save('Email_Analytics.xlsx')
    print("Write successful....")

def delete(imapobj, email_data):
    imapobj.select_folder('Inbox', readonly=False)
    delete_count = 0
    for i in range(len(UIDs)):
        if 'Best Buy' in str(email_data[5][i][0]):
            # For Gmail
            imapobj.add_gmail_labels(UIDs[i],'\Trash')

            # For other clients
            # imapobj.delete_messages(UIDs[i])
            # imapobj.expunge()
            delete_count += 1

    print(str(delete_count) + "emails successfully deleted....")

def unsubscribe(smtpobj, email_data, username):
    for i in range(len(UIDs)):
        if 'Groupon' in str(email_data[5][i][0]):
            unsub_link = email_data[9][i]
            break
    if unsub_link == 'No unsubscribe link found':
        print ("Unsubscribe failed. No link found...")
    else:
        unsub_link = re.sub('[<>]', '', unsub_link)
        parsed = urllib.parse.urlparse(unsub_link)
        unsub_address = parsed.path
        fields = urllib.parse.parse_qs(parsed.query)
        if 'subject' in fields.keys():
            unsub_subject = fields['subject'][0]
        else:
            unsub_subject = 'Unsubscribe'
        msg = MIMEMultipart()
        msg['To'] = unsub_address
        msg['Subject'] = unsub_subject
        smtpobj.send_message(msg, username, [unsub_address])
        print ("Unsubscribe successful")

def logout(imapobj, smtpobj):
    imapobj.logout()
    smtpobj.quit()

if __name__ == '__main__':
    imap_server = input("Enter IMAP server domain name")
    smtp_server = input("Enter SMTP server domain name")
    username = input("Enter username")
    password = input("Enter password")

    start_time = time.time()

    imapobj, smtpobj = login(imap_server, smtp_server, username, password)
    UIDs = retrieve_emails(imapobj)
    email_data = email_attributes(imapobj, UIDs)
    write_to_excel(email_data)
    end_time = time.time()
    print(str(len(UIDs)) + " emails analyzed")
    print("execution time : " + str((end_time - start_time)))

    delete(imapobj, email_data)
    unsubscribe(smtpobj, email_data, username)





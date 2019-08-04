import imapclient
import pyzmail
import time
import datetime
import calendar
import re
import pyprind

from openpyxl import Workbook
from openpyxl import load_workbook

# starting execution time counter
start_time = time.time()

# login credentials
username = 'karthiek.umich@gmail.com'
#password = input("enter password : ")
password = 'amgmercedesM1596.2'

# specify client and login
imapobj = imapclient.IMAPClient('imap.gmail.com', ssl=True)
imapobj.login(username, password)

# initializing lists to hold various types of information
from_addresses = []
subjects = []
dates = []
days = []
months = []
years = []
times = []
category = []
sent_received = []
unsub_links = []
total_UIDs = 0

# specify search folders
folder_list = ['[Gmail]/All Mail', '[Gmail]/Spam', '[Gmail]/Trash']

for folder in folder_list:
    imapobj.select_folder(folder, readonly=True)
    UIDs = imapobj.search(['SINCE', '02-Aug-2019', 'BEFORE', '09-Aug-2019'])

    current_UIDs = len(UIDs)
    total_UIDs += len(UIDs)

    # progress bar set by number of emails in current folder
    n = current_UIDs
    bar = pyprind.ProgBar(n)

    if folder == '[Gmail]/All Mail':
        print ("Analyzing all folders...")
    elif folder == '[Gmail]/Spam':
        print("Analyzing spam...")
    elif folder == '[Gmail]/Trash':
        print("Analyzing trash...")

    for i in range(len(UIDs)):

        # get email label
        label_dict = imapobj.get_gmail_labels(UIDs[i])

        # need to clean up using regex
        label = str(label_dict.get(UIDs[i])).strip('[').strip(']').replace('\'','').replace('\\','')

        if folder == '[Gmail]/Spam':
            category.append('Spam')
        elif folder == '[Gmail]/Trash':
            category.append('Trash')
        else:
            if label == 'Inbox':
                category.append('Inbox')
            elif label == 'Sent':
                category.append('Sent Mail')
            elif 'Starred' in label:
                category.append('Starred')
            elif 'Important' in label:
                category.append('Important')
            else:
                category.append('Custom Label')

        raw_message = imapobj.fetch(UIDs[i], ['BODY.PEEK[HEADER]'])
        message = pyzmail.PyzMessage.factory(raw_message[UIDs[i]][b'BODY[HEADER]'])

        # check whether email is a sent message or a received message
        if message.get_address('from')[1] == username:
            sent_received.append('Sent')
        else:
            sent_received.append('Received')

        from_addresses.append(message.get_address('from'))
        subjects.append(message.get_subject(''))
        full_date = message.get_decoded_header('date')

        # two types of date formats are possible
        if (',' in full_date):
            day = full_date.split(', ')[0]
            date = full_date.split(', ')[1].split()[0]
            if date[0] == '0':
                date = date.split()[0][1:]
            month = full_date.split(', ')[1].split()[1]
            year = full_date.split(', ')[1].split()[2]
            time2 = full_date.split(', ')[1].split()[3]
        else:
            date = full_date.split()[0]
            if date[0] == '0':
                date = date.split()[0][1:]
            month = full_date.split()[1]
            year = full_date.split()[2]
            time2 = full_date.split()[3]
            month_no = list(calendar.month_abbr).index(month)
            day = (calendar.day_name[datetime.date(int(year), int(month_no), int(date)).weekday()]).split()[0][0:3]

        days.append(day)
        dates.append(date)
        months.append(month)
        years.append(year)
        times.append(time2)

        unsub_link = message.get_decoded_header('List-Unsubscribe')
        if len(str(unsub_link)) > 0:
            unsub_links.append(unsub_link)
        else:
            unsub_links.append('No unsubscribe link found')

        # update progress bar after a small delay
        time.sleep(0.01)
        bar.update()

# logout of client
imapobj.logout()

# write extracted information to an excel sheet
wb = Workbook()
ws = wb.active
ws.title = "Data"
ws.cell(1,1).value = "Date"
ws.cell(1,2).value = "Month"
ws.cell(1,3).value = "Year"
ws.cell(1,4).value = "Day"
ws.cell(1,5).value = "Time"
ws.cell(1,6).value = "From (Sender)"
ws.cell(1,7).value = "From (Email ID)"
ws.cell(1,8).value = "Subject"
ws.cell(1,9).value = "Sent/Received"
ws.cell(1,10).value = "Category"
ws.cell(1,11).value = "Unsubscribe Link"

for i in range(total_UIDs):
    ws.cell(row=i+2, column=1).value = dates[i]
    ws.cell(row=i+2, column=2).value = months[i]
    ws.cell(row=i+2, column=3).value = years[i]
    ws.cell(row=i+2, column=4).value = days[i]
    ws.cell(row=i+2, column=5).value = times[i]
    ws.cell(row = i+2, column = 6).value = from_addresses[i][0]
    ws.cell(row = i+2, column = 7).value = from_addresses[i][1]
    ws.cell(row = i+2, column = 8).value = str(subjects[i])
    ws.cell(row=i + 2, column=9).value = sent_received[i]
    ws.cell(row=i+2, column=10).value = category[i]
    ws.cell(row=i + 2, column=11).value = unsub_links[i]

wb.save('Email_Analytics.xlsx')

# ending execution time counter
end_time = time.time()
print ("\n")
print (str(total_UIDs) + " emails analyzed")
print ("execution time : " + str((end_time-start_time)))

